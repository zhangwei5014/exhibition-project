"""
EPMS - 展览项目管理系统
Streamlit Application
"""
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, date
import sqlite3
import hashlib
import io
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import os

# ============================================================
# 配置
# ============================================================
DB_PATH = "epms.db"
DEMO_PASSWORD_HASH = hashlib.sha256("admin123".encode()).hexdigest()

# ============================================================
# 数据库初始化
# ============================================================
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT DEFAULT 'member',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            description TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            phase TEXT,
            name TEXT NOT NULL,
            due_date DATE,
            status TEXT DEFAULT 'pending',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects(id)
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS daily_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            project_id INTEGER,
            report_date DATE,
            content TEXT,
            problems TEXT,
            next_plan TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (user_id) REFERENCES users(id),
            FOREIGN KEY (project_id) REFERENCES projects(id)
        )
    """)
    # 添加默认管理员
    c.execute("SELECT COUNT(*) FROM users")
    if c.fetchone()[0] == 0:
        c.execute("INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
                  ("admin", DEMO_PASSWORD_HASH, "admin"))
        c.execute("INSERT INTO projects (name, description) VALUES (?, ?)",
                  ("江苏移动全业务展厅项目", "江苏移动全业务展厅施工进度管理"))
    conn.commit()
    conn.close()

# ============================================================
# 认证
# ============================================================
def check_auth():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
        st.session_state.username = None
        st.session_state.role = None

def login_user(username, password):
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    password_hash = hashlib.sha256(password.encode()).hexdigest()
    c.execute("SELECT id, username, role FROM users WHERE username=? AND password_hash=?",
              (username, password_hash))
    user = c.fetchone()
    conn.close()
    return user

def render_login():
    st.set_page_config(page_title="EPMS - 登录", page_icon="📋")
    st.title("🏗️ EPMS - 展览项目管理系统")
    st.markdown("---")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("### 👤 用户登录")
        username = st.text_input("用户名", placeholder="请输入用户名")
        password = st.text_input("密码", type="password", placeholder="请输入密码")
        
        if st.button("登录", use_container_width=True):
            if username and password:
                user = login_user(username, password)
                if user:
                    st.session_state.logged_in = True
                    st.session_state.username = user[1]
                    st.session_state.role = user[2]
                    st.rerun()
                else:
                    st.error("❌ 用户名或密码错误")
            else:
                st.warning("请输入用户名和密码")
        
        st.markdown("---")
        st.caption("默认账号: admin / admin123")

# ============================================================
# Excel 导入
# ============================================================
def parse_excel_template(file_bytes):
    """解析Excel模板，返回任务列表"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    
    tasks = []
    for row in ws.iter_rows(min_row=8, values_only=True):
        phase = row[1]  # B列：工作阶段
        seq = row[2]    # C列：序号
        name = row[3]   # D列：主要工作内容
        due_date = row[4]  # E列：完成日期
        
        if name and isinstance(due_date, datetime):
            tasks.append({
                "phase": phase or "未分类",
                "seq": seq,
                "name": name,
                "due_date": due_date.date(),
                "status": "pending"
            })
    
    return tasks

def generate_excel_template(tasks):
    """生成Excel（对齐模板格式）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "实施进度表"
    
    # 标题行
    ws.merge_cells("A1:G1")
    ws["A1"] = "江苏移动全业务展厅项目-进度计划表"
    ws["A1"].font = Font(bold=True, size=14)
    
    # 图例
    ws["F3"] = "红色表示实施重要节点"
    ws["F4"] = "黄色表示施工板块"
    ws["F5"] = "绿色表示内容实施板块"
    
    # 更新日期
    ws["A6"] = "update:"
    ws["B6"] = "=NOW()"
    
    # 表头
    headers = ["工作阶段", "工作阶段", "序号", "主要工作内容", "完成日期", "倒计时天"]
    for i, h in enumerate(headers, 1):
        ws.cell(7, i, h)
    
    # 填充数据
    red_fill = PatternFill("solid", fgColor="FFCCCC")
    yellow_fill = PatternFill("solid", fgColor="FFFF99")
    green_fill = PatternFill("solid", fgColor="CCFFCC")
    
    for i, task in enumerate(tasks, 8):
        ws.cell(i, 2, task.get("phase", ""))
        ws.cell(i, 3, task.get("seq", ""))
        ws.cell(i, 4, task.get("name", ""))
        ws.cell(i, 5, task.get("due_date", ""))
        
        # 倒计时公式
        ws.cell(i, 6, f'=IF(TODAY()>E{i},"已过期",IF(DATEDIF(TODAY(),E{i},"d")<1,"已过期",DATEDIF(TODAY(),E{i},"d")))')
        
        # 根据阶段着色
        phase = task.get("phase", "")
        if "设计" in str(phase) or "商务" in str(phase):
            fill = red_fill
        elif "施工" in str(phase) or "准备" in str(phase):
            fill = yellow_fill
        else:
            fill = green_fill
        
        for col in range(1, 7):
            ws.cell(i, col).fill = fill
    
    wb.save("进度计划表_导出.xlsx")
    return "进度计划表_导出.xlsx"

# ============================================================
# 主界面
# ============================================================
def render_main():
    st.set_page_config(
        page_title="EPMS - 展览项目管理系统",
        page_icon="🏗️",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # 侧边栏
    with st.sidebar:
        st.markdown("### 🏗️ EPMS")
        st.markdown(f"**👤 {st.session_state.username}**")
        st.markdown("---")
        
        page = st.radio("功能导航", [
            "📊 任务看板",
            "📅 风险预警",
            "📝 日报管理",
            "📁 Excel导入/导出",
            "👥 团队管理",
        ], index=0)
        
        st.markdown("---")
        if st.button("🚪 退出登录"):
            st.session_state.logged_in = False
            st.session_state.username = None
            st.rerun()
    
    if page == "📊 任务看板":
        render_task_board()
    elif page == "📅 风险预警":
        render_risk_warning()
    elif page == "📝 日报管理":
        render_daily_report()
    elif page == "📁 Excel导入/导出":
        render_excel_page()
    elif page == "👥 团队管理":
        render_team_page()

# ============================================================
# 任务看板
# ============================================================
def render_task_board():
    st.title("📊 任务看板")
    
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql("""
        SELECT t.*, p.name as project_name 
        FROM tasks t 
        LEFT JOIN projects p ON t.project_id = p.id
        ORDER BY t.due_date
    """, conn, parse_dates=["due_date"])
    conn.close()
    
    # 筛选器
    col1, col2, col3 = st.columns(3)
    with col1:
        status_filter = st.selectbox("状态", ["全部", "pending", "进行中", "已完成"])
    with col2:
        phase_filter = st.selectbox("阶段", ["全部"] + list(df["phase"].dropna().unique()))
    with col3:
        search = st.text_input("搜索任务", "")
    
    # 应用筛选
    if status_filter != "全部":
        df = df[df["status"] == status_filter]
    if phase_filter != "全部":
        df = df[df["phase"] == phase_filter]
    if search:
        df = df[df["name"].str.contains(search, na=False)]
    
    # 统计卡片
    today = date.today()
    total = len(df)
    overdue = len(df[df["due_date"].dt.date < today])
    soon = len(df[(df["due_date"].dt.date - today).dt.days <= 3])
    completed = len(df[df["status"] == "completed"])
    
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("总任务", total)
    c2.metric("已完成", completed, delta=f"{completed*100//max(total,1)}%" if total else 0)
    c3.metric("⚠️ 即将到期(3天内)", soon)
    c4.metric("🔴 已逾期", overdue)
    
    st.markdown("---")
    
    # 按阶段分组显示
    phases = df["phase"].dropna().unique()
    for phase in phases:
        with st.expander(f"**{phase}** ({len(df[df['phase']==phase])})", expanded=True):
            phase_df = df[df["phase"] == phase]
            cols = st.columns([1, 3, 1, 1, 1])
            cols[0].write("**序号**")
            cols[1].write("**任务名称**")
            cols[2].write("**截止日期**")
            cols[3].write("**状态**")
            cols[4].write("**操作**")
            
            for _, row in phase_df.iterrows():
                cols = st.columns([1, 3, 1, 1, 1])
                cols[0].write(str(row["seq"]) if pd.notna(row["seq"]) else "-")
                cols[1].write(row["name"])
                
                due = row["due_date"]
                if pd.notna(due):
                    days_left = (due.date() - today).days
                    if days_left < 0:
                        cols[2].markdown(f"🔴 **{due.strftime('%m-%d')}** (已逾期{-days_left}天)")
                    elif days_left <= 3:
                        cols[2].markdown(f"🟡 **{due.strftime('%m-%d')}** ({days_left}天)")
                    else:
                        cols[2].write(due.strftime("%m-%d"))
                else:
                    cols[2].write("-")
                
                status_emoji = {"pending": "⏳", "in_progress": "🔄", "completed": "✅"}.get(row["status"], "⏳")
                cols[3].write(f"{status_emoji} {row['status']}")
                
                with cols[4]:
                    status_map = {"pending": 0, "in_progress": 1, "completed": 2}
                    new_statuses = ["pending", "in_progress", "completed"]
                    idx = status_map.get(row["status"], 0)
                    new_status = st.selectbox(
                        "更新状态", 
                        new_statuses,
                        index=idx,
                        key=f"status_{row['id']}",
                        label_visibility="collapsed"
                    )
                    if new_status != row["status"]:
                        conn2 = sqlite3.connect(DB_PATH)
                        conn2.execute("UPDATE tasks SET status=? WHERE id=?", (new_status, row["id"]))
                        conn2.commit()
                        conn2.close()
                        st.rerun()

# ============================================================
# 风险预警
# ============================================================
def render_risk_warning():
    st.title("📅 风险预警")
    
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql("""
        SELECT t.*, p.name as project_name 
        FROM tasks t 
        LEFT JOIN projects p ON t.project_id = p.id
        ORDER BY t.due_date
    """, conn, parse_dates=["due_date"])
    conn.close()
    
    today = date.today()
    df["days_left"] = df["due_date"].apply(lambda x: (x.date() - today).days if pd.notna(x) else None)
    
    # 甘特图
    st.markdown("### 📊 甘特图")
    active_df = df[(df["status"] != "completed") & df["due_date"].notna()]
    if not active_df.empty:
        fig = px.timeline(
            active_df,
            x_start="due_date",
            x_end="due_date",
            y="phase",
            color="phase",
            hover_name="name",
            text="name"
        )
        fig.update_layout(showlegend=False, height=400)
        st.plotly_chart(fig, use_container_width=True)
    
    # 预警列表
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("### 🔴 已逾期任务")
        overdue_df = df[df["days_left"] < 0]
        if overdue_df.empty:
            st.success("✅ 暂无逾期任务")
        else:
            for _, row in overdue_df.iterrows():
                st.error(f"**{row['name']}** (逾期{-row['days_left']}天) - {row['phase']}")
    
    with col2:
        st.markdown("### 🟡 即将到期(3天内)")
        soon_df = df[(df["days_left"] >= 0) & (df["days_left"] <= 3)]
        if soon_df.empty:
            st.success("✅ 近期无紧急任务")
        else:
            for _, row in soon_df.iterrows():
                st.warning(f"**{row['name']}** ({row['days_left']}天后到期) - {row['phase']}")

# ============================================================
# 日报管理
# ============================================================
def render_daily_report():
    st.title("📝 施工日报管理")
    
    tab1, tab2 = st.tabs(["✍️ 填写日报", "📋 历史记录"])
    
    with tab1:
        with st.form("daily_report_form"):
            st.markdown("#### 填写施工日报")
            c1, c2 = st.columns(2)
            with c1:
                report_date = st.date_input("日期", value=date.today())
            with c2:
                project = st.selectbox("选择项目", ["江苏移动全业务展厅项目"])
            
            content = st.text_area("📌 今日工作内容", height=100, placeholder="描述今日完成的工作...")
            problems = st.text_area("⚠️ 遇到的问题", height=80, placeholder="描述遇到的问题（如无则留空）...")
            next_plan = st.text_area("📅 次日工作计划", height=80, placeholder="描述次日计划...")
            
            if st.form_submit_button("📤 提交日报", use_container_width=True):
                if content:
                    conn = sqlite3.connect(DB_PATH)
                    c = conn.cursor()
                    c.execute("SELECT id FROM users WHERE username=?", (st.session_state.username,))
                    user_id = c.fetchone()[0]
                    c.execute("SELECT id FROM projects LIMIT 1")
                    project_id = c.fetchone()[0]
                    c.execute("""
                        INSERT INTO daily_reports (user_id, project_id, report_date, content, problems, next_plan)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (user_id, project_id, report_date, content, problems, next_plan))
                    conn.commit()
                    conn.close()
                    st.success("✅ 日报提交成功！")
                else:
                    st.error("请填写工作内容")
    
    with tab2:
        st.markdown("#### 历史日报")
        conn = sqlite3.connect(DB_PATH)
        df = pd.read_sql("""
            SELECT r.*, u.username, p.name as project_name 
            FROM daily_reports r
            JOIN users u ON r.user_id = u.id
            JOIN projects p ON r.project_id = p.id
            ORDER BY r.report_date DESC
            LIMIT 50
        """, conn, parse_dates=["report_date"])
        conn.close()
        
        for _, row in df.iterrows():
            with st.expander(f"📅 {row['report_date'].strftime('%Y-%m-%d')} - {row['username']}"):
                st.markdown(f"**项目：** {row['project_name']}")
                st.markdown(f"**工作内容：**\n{row['content']}")
                if row['problems']:
                    st.markdown(f"**遇到的问题：** {row['problems']}")
                if row['next_plan']:
                    st.markdown(f"**次日计划：** {row['next_plan']}")

# ============================================================
# Excel 导入/导出
# ============================================================
def render_excel_page():
    st.title("📁 Excel 导入/导出")
    
    tab1, tab2 = st.tabs(["⬆️ 从Excel导入", "⬇️ 导出到Excel"])
    
    with tab1:
        st.markdown("#### 导入 Excel 进度表")
        st.info("支持从Excel模板导入任务数据。导入前请确保格式与模板一致。")
        
        uploaded = st.file_uploader("选择 Excel 文件", type=["xlsx", "xls"], key="import")
        if uploaded:
            if st.button("🚀 开始导入", use_container_width=True):
                with st.spinner("正在解析文件..."):
                    tasks = parse_excel_template(uploaded.read())
                    if tasks:
                        conn = sqlite3.connect(DB_PATH)
                        c = conn.cursor()
                        c.execute("SELECT id FROM projects LIMIT 1")
                        project_id = c.fetchone()[0]
                        
                        imported = 0
                        for task in tasks:
                            c.execute("""
                                INSERT INTO tasks (project_id, phase, name, due_date, status)
                                VALUES (?, ?, ?, ?, ?)
                            """, (project_id, task["phase"], task["name"], task["due_date"], "pending"))
                            imported += 1
                        
                        conn.commit()
                        conn.close()
                        st.success(f"✅ 成功导入 {imported} 个任务！")
                    else:
                        st.error("未找到有效任务数据，请检查文件格式。")
    
    with tab2:
        st.markdown("#### 导出为 Excel")
        st.info("将当前所有任务导出为标准Excel格式，保持与模板一致的列结构。")
        
        if st.button("📥 导出任务列表", use_container_width=True):
            conn = sqlite3.connect(DB_PATH)
            df = pd.read_sql("""
                SELECT phase as '工作阶段', seq as '序号', name as '主要工作内容', 
                       due_date as '完成日期', status as '状态' 
                FROM tasks ORDER BY due_date
            """, conn)
            conn.close()
            
            if not df.empty:
                tasks = df.to_dict("records")
                filepath = generate_excel_template(tasks)
                with open(filepath, "rb") as f:
                    st.download_button(
                        "⬇️ 下载 Excel",
                        f.read(),
                        filepath,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                st.success("✅ 导出成功！")
            else:
                st.warning("暂无任务数据可导出。")

# ============================================================
# 团队管理
# ============================================================
def render_team_page():
    st.title("👥 团队管理")
    
    if st.session_state.role != "admin":
        st.warning("只有管理员可以管理团队成员。")
        return
    
    tab1, tab2 = st.tabs(["👤 成员列表", "➕ 添加成员"])
    
    with tab1:
        conn = sqlite3.connect(DB_PATH)
        df = pd.read_sql("SELECT id, username, role, created_at FROM users", conn)
        conn.close()
        st.dataframe(df, use_container_width=True)
    
    with tab2:
        with st.form("add_user_form"):
            new_username = st.text_input("用户名")
            new_password = st.text_input("密码", type="password")
            new_role = st.selectbox("角色", ["member", "admin"])
            if st.form_submit_button("添加成员"):
                if new_username and new_password:
                    try:
                        conn = sqlite3.connect(DB_PATH)
                        c = conn.cursor()
                        c.execute("INSERT INTO users (username, password_hash, role) VALUES (?, ?, ?)",
                                  (new_username, hashlib.sha256(new_password.encode()).hexdigest(), new_role))
                        conn.commit()
                        conn.close()
                        st.success("✅ 成员添加成功！")
                    except:
                        st.error("用户名已存在！")
                else:
                    st.error("请填写完整信息")

# ============================================================
# 启动
# ============================================================
if __name__ == "__main__":
    init_db()
    check_auth()
    
    if not st.session_state.logged_in:
        render_login()
    else:
        render_main()
